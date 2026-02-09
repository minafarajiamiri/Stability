# -*- coding: utf-8 -*-
"""
Created on Sun Feb  8 12:03:37 2026

@author: amiri

Fill Analysis_MasterTemplate.xlsx -> Analysis_MasterTemplate_filled_correctIndex.xlsx

The script:
1) Reads analysis1_input from the template
2) Loads all results_*.json files
3) Builds a lookup keyed by (dataset, method, model, excel_question_id)
4) Cross-checks and corrects predicted_answer and is_correct
5) Writes back the updated sheet and prints a discrepancy summary

(Answer extraction):
- Parses item["predicted_answer"] to robustly extract the FINAL chosen option (A–E)
- Prioritizes the LAST occurrence of any "heavy phrase" markers
- Falls back to explicit option letters if no heavy phrases found
- Falls back to semantic match against option texts if needed
"""

from __future__ import annotations

import os
import glob
import json
import re
import string
import difflib
from typing import Any, Dict, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# -----------------------------
# CONFIG
# -----------------------------
TEMPLATE_XLSX = "Analysis_MasterTemplate.xlsx"
OUTPUT_XLSX = "Analysis_MasterTemplate_filled_correctIndex_V2.xlsx"

# IMPORTANT: JSON ids start at 0, Excel ids start at 1
JSON_QID_STARTS_AT_ZERO = True

TARGET_SHEET = "analysis1_input"

# -----------------------------
# Parsing helpers (A–E)
# -----------------------------
STOP = set(
    """
    the a an and or of to in on for with without is are was were be been being as at by from
    this that these those it its their her his patient presents presenting findings show shows
    consistent most likely diagnosis
    """.split()
)

def normalize_words(s: str) -> list[str]:
    s = (s or "").lower()
    s = s.translate(str.maketrans({c: " " for c in string.punctuation}))
    return [w for w in s.split() if w and w not in STOP]


def infer_choice_from_options(pred_text: str, options: dict) -> Optional[str]:
    """Fallback: match free-text prediction to option texts."""
    if not pred_text or not isinstance(options, dict) or not options:
        return None

    pw = set(normalize_words(pred_text))
    if not pw:
        return None

    best, best_score = None, 0.0
    for k, opt_text in options.items():
        if not opt_text:
            continue

        ow = set(normalize_words(str(opt_text)))
        if not ow:
            continue

        j = len(pw & ow) / len(pw | ow) if (pw | ow) else 0.0
        seq = difflib.SequenceMatcher(None, pred_text.lower(), str(opt_text).lower()).ratio()
        score = 0.7 * j + 0.3 * seq

        if str(opt_text).lower() in pred_text.lower():
            score += 0.5

        if score > best_score:
            best_score, best = score, k

    return str(best).strip().upper() if best_score >= 0.08 else None


# -----------------------------
# NEW: robust extraction from predicted_answer
# -----------------------------
HEAVY_PHRASES = [
    r"the\s+final\s+answer\s+is\s*:",
    r"\*\*\s*Option\s+or\s+the\s+correct\s+answer\s+is\s*\*\*",
    r"therefore,\s+the\s+answer\s+is\s*:",
    r"\*\*\s*Option\s+or\s+Final\s+Answer\s*:\s*\*\*",
    r"the\s+best\s+answer\s+is\s*:",
    r"\*\*\s*Option\s+or\s+appropriate\s+answer\s+is\s*:\s*\*\*",
]

HEAVY_PHRASE_RE = re.compile(
    "(" + "|".join(HEAVY_PHRASES) + ")",
    flags=re.IGNORECASE,
)

OPTION_AFTER_PHRASE_RE = re.compile(
    r"""
    ^\s*
    (?:\*{0,2}\s*)?
    (?:option\s*)?
    [\(\[]?\s*
    (?P<opt>[A-E])
    \s*[\)\]]?
    \s*(?:[\:\.\)\-]|$|\b)
    """,
    flags=re.IGNORECASE | re.VERBOSE,
)

FALLBACK_OPTION_PATTERNS = [
    r"\bfinal\s*answer\s*[:\-]?\s*\*{0,2}\s*([A-E])\b",
    r"\btherefore\s*,?\s*the\s+answer\s+is\s*[:\-]?\s*\*{0,2}\s*([A-E])\b",
    r"\bthe\s+best\s+answer\s+is\s*[:\-]?\s*\*{0,2}\s*([A-E])\b",
    r"(?m)^\s*([A-E])\s*[:\.\)]\s*",
    r"\*\*\s*([A-E])\s*\*\*",
    r"\boption\s+([A-E])\b",
    r"\banswer\s*[:\-]?\s*\*{0,2}\s*([A-E])\b",
]

FALLBACK_OPTION_RE_LIST = [
    re.compile(p, flags=re.IGNORECASE | re.DOTALL)
    for p in FALLBACK_OPTION_PATTERNS
]


def extract_final_option(predicted_answer: Any) -> Optional[str]:
    """
    Extract the final chosen option (A–E) from predicted_answer.

    Priority:
      1) Find any heavy phrase; if multiple exist, use LAST occurrence.
         Then extract the option immediately following that phrase.
      2) If no heavy phrase found, scan whole text for explicit option patterns
         (use LAST occurrence of matched pattern).
      3) If still nothing, return None.
    """
    if predicted_answer is None:
        return None

    t = str(predicted_answer).strip()
    if not t:
        return None

    if t in list("ABCDE"):
        return t

    # 1) Heavy phrase search -> LAST occurrence
    matches = list(HEAVY_PHRASE_RE.finditer(t))
    if matches:
        last = matches[-1]
        tail = t[last.end():]  # substring after the LAST phrase

        m = OPTION_AFTER_PHRASE_RE.search(tail)
        if m:
            return m.group("opt").upper()

        # If not immediately after phrase, search tail for standalone letter
        m2 = re.search(r"\b([A-E])\b", tail, flags=re.IGNORECASE)
        if m2:
            return m2.group(1).upper()

    # 2) Fallback patterns on whole text (also use LAST occurrence)
    for rx in FALLBACK_OPTION_RE_LIST:
        found = list(rx.finditer(t))
        if found:
            return found[-1].group(1).upper()

    return None


def extract_choice_with_fallback(item: dict) -> Optional[str]:
    c = extract_final_option(item.get("predicted_answer", ""))
    if c:
        return c
    return infer_choice_from_options(item.get("predicted_answer", ""), item.get("options", {}))


# -----------------------------
# Filename → dataset/method/model
# -----------------------------
MODEL_MAP = {
    "Ministral-8B-Instruct-2410": "Ministral-8B",
    "mistrallarge123b": "Mistral Large (123B)",
    "Meta-Llama-3-8B-Instruct": "Llama3.3-8B",
    "llama3_370b": "Llama3.3-70B",
    "Llama3-Med42-8B": "Llama3-Med42-8B",
    "Llama3-Med42-70B": "Llama3-Med42-70B",
    "llama4scout16E": "Llama4 Scout 16E",
    "deepseek70br1": "DeepSeek R1-70B",
    "deepseekr1": "DeepSeek-R1 (671B)",
    "DeepSeek-V3": "DeepSeek-V3 (671B)",
    "Qwen2.5-0.5B-Instruct": "Qwen 2.5-0.5B",
    "Qwen2.5-3B-Instruct": "Qwen 2.5-3B",
    "Qwen2.5-7B-Instruct": "Qwen 2.5-7B",
    "Qwen2.5-14B-Instruct": "Qwen 2.5-14B",
    "qwen2570bins": "Qwen 2.5-70B",
    "Qwen3-8B": "Qwen 3-8B",
    "Qwen3-235B-A22B-Instruct-2507": "Qwen 3-235B",
    "gemma-3-4b-it": "Gemma-3-4B-it",
    "gemma-3-27b-it": "Gemma-3-27B-it",
    "medgemma-4b-it": "MedGemma-4B-it",
    "medgemma-27b-text-it": "MedGemma-27B-text-it",
    "gpt-3.5-turbo": "GPT-3.5-turbo",
    "gpt-4-turbo": "GPT-4-turbo",
    "gpt-5": "GPT-5",
    "o3": "o3",
}

def parse_filename(fp: str) -> Optional[Tuple[str, str, str]]:
    """
    Supports:
      results_agentic_internal_dataset_dr_<MODEL>.json
      results_zeroshot_radiology_dr_final_<MODEL>.json
    Returns (method, dataset, model_display)
    """
    bn = os.path.basename(fp)
    m = re.match(
        r"results_(agentic|zeroshot)_(internal_dataset|radiology)_dr(?:_final)?_(.+)\.json$",
        bn,
        flags=re.IGNORECASE,
    )
    if not m:
        return None

    method_raw, dataset_raw, model_raw = m.group(1), m.group(2), m.group(3)
    method = "agentic" if method_raw.lower() == "agentic" else "zero-shot"
    dataset = "Internal_TUM" if dataset_raw.lower() == "internal_dataset" else "RadioRAG"
    model = MODEL_MAP.get(model_raw, model_raw)
    return method, dataset, model


# -----------------------------
# Excel styling helper
# -----------------------------
def style_as_table(ws, table_name: str):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    for i in range(1, ws.max_column + 1):
        maxlen = len(str(ws.cell(1, i).value or ""))
        for rr in range(2, min(ws.max_row, 200) + 1):
            v = ws.cell(rr, i).value
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, maxlen + 2), 60)

    tab = Table(displayName=table_name, ref=ws.dimensions)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)


# -----------------------------
# Main
# -----------------------------
def main():
    if not os.path.exists(TEMPLATE_XLSX):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_XLSX}")

    wb = load_workbook(TEMPLATE_XLSX)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TARGET_SHEET}' not found in template.")

    ws = wb[TARGET_SHEET]
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    df = pd.DataFrame(rows)

    required = {"dataset", "method", "model", "question_id", "correct_answer"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{TARGET_SHEET} missing required columns: {sorted(missing)}")

    # --- Build JSON lookup keyed by (dataset, method, model, excel_qid) ---
    lookup: Dict[Tuple[str, str, str, int], Tuple[Optional[str], Optional[int], str]] = {}

    json_files = glob.glob("results_*_*.json")
    if not json_files:
        print("WARNING: No JSON files found (pattern results_*_*.json).")

    for fp in sorted(json_files):
        meta = parse_filename(fp)
        if meta is None:
            continue
        method, dataset, model = meta

        try:
            with open(fp, "r", encoding="utf-8") as f:
                items = json.load(f)
        except Exception as e:
            print(f"[SKIP] {fp}: JSON error: {e}")
            continue

        if not isinstance(items, list):
            print(f"[SKIP] {fp}: top-level JSON is not a list")
            continue

        for it in items:
            if not isinstance(it, dict) or "question_id" not in it:
                continue

            try:
                qid_json = int(it["question_id"])
            except Exception:
                continue

            # *** CRITICAL FIX: JSON 0-based → Excel 1-based ***
            qid_excel = qid_json + 1 if JSON_QID_STARTS_AT_ZERO else qid_json

            pred = extract_choice_with_fallback(it)

            corr = it.get("correct")
            if corr is None:
                # If 'correct' missing, compute via answer_idx when possible
                gt = it.get("answer_idx")
                if gt is not None and pred is not None:
                    corr = int(str(pred) == str(gt))

            try:
                corr_int = int(corr) if corr is not None else None
            except Exception:
                corr_int = None

            lookup[(dataset, method, model, qid_excel)] = (pred, corr_int, os.path.basename(fp))

    # --- Cross-check + correct Excel columns ---
    df["orig_predicted_answer"] = df.get("predicted_answer")
    df["orig_is_correct"] = df.get("is_correct")

    new_pred, new_corr, src_file, match_status = [], [], [], []

    for _, r in df.iterrows():
        key = (r["dataset"], r["method"], r["model"], int(r["question_id"]))
        if key in lookup:
            p, c, src = lookup[key]
            new_pred.append(p)
            new_corr.append(c)
            src_file.append(src)
            match_status.append("matched")
        else:
            new_pred.append(np.nan)
            new_corr.append(np.nan)
            src_file.append(np.nan)
            match_status.append("no_json_match")

    df["predicted_answer"] = new_pred
    df["is_correct"] = new_corr
    df["prediction_source_file"] = src_file
    df["json_match_status"] = match_status

    # --- Discrepancy summary (before vs after) ---
    matched_mask = df["json_match_status"] == "matched"

    pred_changed = (
        (df["orig_predicted_answer"].astype(str) != df["predicted_answer"].astype(str))
        & matched_mask
    ).sum()

    corr_changed = (
        (pd.to_numeric(df["orig_is_correct"], errors="coerce")
         != pd.to_numeric(df["is_correct"], errors="coerce"))
        & matched_mask
    ).sum()

    total = len(df)
    matched = int(matched_mask.sum())
    no_match = total - matched

    # --- Write back to workbook (replace sheet content) ---
    del wb[TARGET_SHEET]
    ws2 = wb.create_sheet(TARGET_SHEET, 0)

    ws2.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws2.append(list(row))

    style_as_table(ws2, "tblAnalysis1Input")
    wb.save(OUTPUT_XLSX)

    print("\n=== SUMMARY ===")
    print(f"Output written: {OUTPUT_XLSX}")
    print(f"Total rows: {total}")
    print(f"Matched to JSON (after correct indexing): {matched}")
    print(f"No JSON match: {no_match}")
    print(f"Rows corrected (predicted_answer): {int(pred_changed)}")
    print(f"Rows corrected (is_correct): {int(corr_changed)}")
    print("Indexing handled as: excel_qid = json_qid + 1")


if __name__ == "__main__":
    main()